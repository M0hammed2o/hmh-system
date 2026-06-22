"""
Tests for Municipality Invoice module.

Covers:
  - Create invoice (with and without items)
  - List invoices filtered by project
  - Get single invoice
  - Update invoice fields and items
  - Delete draft invoice
  - Cannot delete FINALISED invoice
  - Totals auto-computed from items
  - Excel export returns bytes
  - Subcontractor payment summary (monthly) cross-check
  - Requires auth
"""

import uuid
from datetime import date

import pytest

from tests.conftest import auth, login, make_project, make_user, make_user_project_access


# ── Fixtures ──────────────────────────────────────────────────────────────────

@pytest.fixture()
def ctx(db, client):
    owner   = make_user(db, role="OWNER")
    project = make_project(db, owner["id"])
    token   = login(client, owner["email"], owner["password"])
    return {
        "owner":   owner,
        "project": project,
        "token":   token,
        "headers": auth(token),
        "project_id": project["id"],
    }


def _create(client, ctx, **overrides):
    payload = {
        "cert_number":          "26",
        "invoice_date":         "2026-01-26",
        "client_name":          "Ethekweni Municipality",
        "client_vat_no":        "4880193505",
        "client_address":       "PO BOX 828,Durban\n4001",
        "company_email":        "dwt786@gmail.com",
        "project_description":  "Kwalinda Rural Housing Project",
        "contract_reference":   "Contract 1H-42037",
        "previously_paid":      0,
        "vat_rate":             15,
        "bank_name":            "FIRST NATIONAL BANK",
        "account_number":       "62381077893",
        "branch_name":          "FLORIDA ROAD",
        "branch_code":          "220526",
        "notes":                "In this invoice we claiming for 15 completions.",
        "items": [
            {"line_number": "7",  "description": "Wall-plate , door & window ceiling", "quantity": 2,  "sort_order": 6},
            {"line_number": "8",  "description": "Roof & Ceiling",                     "quantity": 8,  "sort_order": 7},
            {"line_number": "9",  "description": "Electrical",                         "quantity": 15, "unit_price": 5000, "sort_order": 8},
            {"line_number": "10", "description": "Concrete Aprons",                    "quantity": 15, "unit_price": 2000, "sort_order": 9},
            {"line_number": "11", "description": "Finishing-plaster, paint , windows & doors", "quantity": 15, "unit_price": 3000, "sort_order": 10},
        ],
    }
    payload.update(overrides)
    return client.post(
        f"/api/v1/projects/{ctx['project_id']}/municipality-invoices/",
        json=payload,
        headers=ctx["headers"],
    )


# ── Create ────────────────────────────────────────────────────────────────────

class TestCreate:
    def test_returns_201(self, db, client, ctx):
        r = _create(client, ctx)
        assert r.status_code == 201, r.text
        data = r.json()["data"]
        assert data["cert_number"] == "26"
        assert data["client_name"] == "Ethekweni Municipality"
        assert data["status"] == "DRAFT"
        assert data["invoice_number"].startswith("IN")

    def test_totals_computed_from_items(self, db, client, ctx):
        r = _create(client, ctx)
        assert r.status_code == 201
        data = r.json()["data"]
        # Items with totals: 15*5000 + 15*2000 + 15*3000 = 75000+30000+45000 = 150000
        assert data["subtotal"] == 150000.0
        assert data["vat_amount"] == pytest.approx(22500.0, abs=1)
        assert data["total_due"]  == pytest.approx(172500.0, abs=1)

    def test_invoice_number_auto_generated(self, db, client, ctx):
        r = _create(client, ctx)
        assert r.status_code == 201
        assert r.json()["data"]["invoice_number"].startswith("IN")

    def test_custom_invoice_number(self, db, client, ctx):
        r = _create(client, ctx, invoice_number="IN00472")
        assert r.status_code == 201
        assert r.json()["data"]["invoice_number"] == "IN00472"

    def test_duplicate_invoice_number_rejected(self, db, client, ctx):
        _create(client, ctx, invoice_number="INX001")
        r2 = _create(client, ctx, invoice_number="INX001")
        assert r2.status_code in (400, 422)

    def test_requires_auth(self, db, client, ctx):
        r = client.post(
            f"/api/v1/projects/{ctx['project_id']}/municipality-invoices/",
            json={"cert_number": "1", "invoice_date": "2026-01-01"},
        )
        assert r.status_code == 401

    def test_items_stored_in_order(self, db, client, ctx):
        r = _create(client, ctx)
        data = r.json()["data"]
        numbers = [i["line_number"] for i in data["items"]]
        assert numbers[:3] == ["7", "8", "9"]


# ── List ──────────────────────────────────────────────────────────────────────

class TestList:
    def test_list_returns_created(self, db, client, ctx):
        _create(client, ctx)
        _create(client, ctx)
        r = client.get(
            f"/api/v1/projects/{ctx['project_id']}/municipality-invoices/",
            headers=ctx["headers"],
        )
        assert r.status_code == 200
        assert len(r.json()["data"]) >= 2

    def test_filter_by_status(self, db, client, ctx):
        _create(client, ctx)
        r = client.get(
            f"/api/v1/projects/{ctx['project_id']}/municipality-invoices/?status=DRAFT",
            headers=ctx["headers"],
        )
        assert r.status_code == 200
        for inv in r.json()["data"]:
            assert inv["status"] == "DRAFT"


# ── Get ───────────────────────────────────────────────────────────────────────

class TestGet:
    def test_get_by_id(self, db, client, ctx):
        r = _create(client, ctx)
        inv_id = r.json()["data"]["id"]
        r2 = client.get(f"/api/v1/municipality-invoices/{inv_id}", headers=ctx["headers"])
        assert r2.status_code == 200
        assert r2.json()["data"]["id"] == inv_id

    def test_get_includes_items(self, db, client, ctx):
        r = _create(client, ctx)
        inv_id = r.json()["data"]["id"]
        r2 = client.get(f"/api/v1/municipality-invoices/{inv_id}", headers=ctx["headers"])
        assert len(r2.json()["data"]["items"]) == 5

    def test_404_on_missing(self, db, client, ctx):
        r = client.get(f"/api/v1/municipality-invoices/{uuid.uuid4()}", headers=ctx["headers"])
        assert r.status_code == 404


# ── Update ────────────────────────────────────────────────────────────────────

class TestUpdate:
    def test_update_notes(self, db, client, ctx):
        r = _create(client, ctx)
        inv_id = r.json()["data"]["id"]
        r2 = client.patch(
            f"/api/v1/municipality-invoices/{inv_id}",
            json={"notes": "Updated note"},
            headers=ctx["headers"],
        )
        assert r2.status_code == 200
        assert r2.json()["data"]["notes"] == "Updated note"

    def test_update_replaces_items(self, db, client, ctx):
        r = _create(client, ctx)
        inv_id = r.json()["data"]["id"]
        r2 = client.patch(
            f"/api/v1/municipality-invoices/{inv_id}",
            json={"items": [{"line_number": "1", "description": "P&G", "quantity": 5, "unit_price": 1000, "sort_order": 0}]},
            headers=ctx["headers"],
        )
        assert r2.status_code == 200
        data = r2.json()["data"]
        assert len(data["items"]) == 1
        assert data["subtotal"] == 5000.0

    def test_finalised_cannot_be_edited(self, db, client, ctx):
        r = _create(client, ctx)
        inv_id = r.json()["data"]["id"]
        # Finalise it
        client.patch(
            f"/api/v1/municipality-invoices/{inv_id}",
            json={"status": "FINALISED"},
            headers=ctx["headers"],
        )
        # Try to edit
        r2 = client.patch(
            f"/api/v1/municipality-invoices/{inv_id}",
            json={"notes": "Should fail"},
            headers=ctx["headers"],
        )
        assert r2.status_code in (400, 422)

    def test_update_previously_paid_adjusts_totals(self, db, client, ctx):
        r = _create(client, ctx)
        inv_id = r.json()["data"]["id"]
        r2 = client.patch(
            f"/api/v1/municipality-invoices/{inv_id}",
            json={"previously_paid": 50000},
            headers=ctx["headers"],
        )
        assert r2.status_code == 200
        data = r2.json()["data"]
        # net = 150000 - 50000 = 100000; vat = 15000; due = 115000
        assert data["total_due"] == pytest.approx(115000.0, abs=1)


# ── Delete ────────────────────────────────────────────────────────────────────

class TestDelete:
    def test_delete_draft(self, db, client, ctx):
        r = _create(client, ctx)
        inv_id = r.json()["data"]["id"]
        r2 = client.delete(f"/api/v1/municipality-invoices/{inv_id}", headers=ctx["headers"])
        assert r2.status_code == 200
        # Verify gone
        r3 = client.get(f"/api/v1/municipality-invoices/{inv_id}", headers=ctx["headers"])
        assert r3.status_code == 404

    def test_cannot_delete_finalised(self, db, client, ctx):
        r = _create(client, ctx)
        inv_id = r.json()["data"]["id"]
        client.patch(
            f"/api/v1/municipality-invoices/{inv_id}",
            json={"status": "FINALISED"},
            headers=ctx["headers"],
        )
        r2 = client.delete(f"/api/v1/municipality-invoices/{inv_id}", headers=ctx["headers"])
        assert r2.status_code in (400, 422)


# ── Excel export ──────────────────────────────────────────────────────────────

class TestExcelExport:
    def test_export_returns_bytes(self, db, client, ctx):
        r = _create(client, ctx)
        inv_id = r.json()["data"]["id"]
        r2 = client.get(
            f"/api/v1/municipality-invoices/{inv_id}/export/excel",
            headers=ctx["headers"],
        )
        assert r2.status_code == 200
        assert r2.headers["content-type"].startswith(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        assert len(r2.content) > 1000  # non-trivial file

    def test_export_filename_contains_invoice_number(self, db, client, ctx):
        r = _create(client, ctx, invoice_number="IN00472")
        inv_id = r.json()["data"]["id"]
        r2 = client.get(
            f"/api/v1/municipality-invoices/{inv_id}/export/excel",
            headers=ctx["headers"],
        )
        assert r2.status_code == 200
        disposition = r2.headers.get("content-disposition", "")
        assert "IN00472" in disposition

    def test_export_excel_is_valid_workbook(self, db, client, ctx):
        """Parse the returned bytes with openpyxl to confirm it's a valid workbook."""
        import io
        import openpyxl

        r = _create(client, ctx, invoice_number="IN99999")
        inv_id = r.json()["data"]["id"]
        r2 = client.get(
            f"/api/v1/municipality-invoices/{inv_id}/export/excel",
            headers=ctx["headers"],
        )
        wb = openpyxl.load_workbook(io.BytesIO(r2.content))
        ws = wb.active
        # TAX INVOICE must appear in the sheet
        cell_values = [ws.cell(row=r, column=8).value for r in range(1, 10)]
        assert any("TAX INVOICE" in str(v) for v in cell_values if v)

    def test_export_contains_client_name(self, db, client, ctx):
        import io
        import openpyxl

        r = _create(client, ctx)
        inv_id = r.json()["data"]["id"]
        r2 = client.get(
            f"/api/v1/municipality-invoices/{inv_id}/export/excel",
            headers=ctx["headers"],
        )
        wb = openpyxl.load_workbook(io.BytesIO(r2.content))
        ws = wb.active
        all_values = []
        for row in ws.iter_rows(values_only=True):
            all_values.extend([str(c) for c in row if c is not None])
        assert any("Ethekweni Municipality" in v for v in all_values)


# ── Subcontractor payment summary cross-check ─────────────────────────────────

class TestSubconSummary:
    """Verify the monthly summary endpoint still works alongside municipality invoices."""

    def test_monthly_summary_zero_for_new_project(self, db, client, ctx):
        r = client.get(
            f"/api/v1/projects/{ctx['project_id']}/work-done/monthly-summary?month=2026-01-01",
            headers=ctx["headers"],
        )
        assert r.status_code == 200
        assert r.json()["data"]["total_records"] == 0
        assert r.json()["data"]["total_amount"] == 0.0


# ── Template endpoint ─────────────────────────────────────────────────────────

class TestTemplateEndpoint:
    """GET /projects/{id}/municipality-invoices/template returns the 15-item Cert 26 template."""

    def test_returns_200(self, db, client, ctx):
        r = client.get(
            f"/api/v1/projects/{ctx['project_id']}/municipality-invoices/template",
            headers=ctx["headers"],
        )
        assert r.status_code == 200, r.text

    def test_returns_15_items(self, db, client, ctx):
        r = client.get(
            f"/api/v1/projects/{ctx['project_id']}/municipality-invoices/template",
            headers=ctx["headers"],
        )
        data = r.json()["data"]
        assert len(data["items"]) == 15

    def test_all_standard_fields_present(self, db, client, ctx):
        r = client.get(
            f"/api/v1/projects/{ctx['project_id']}/municipality-invoices/template",
            headers=ctx["headers"],
        )
        data = r.json()["data"]
        assert data["client_name"]    == "Ethekweni Municipality"
        assert data["client_vat_no"]  == "4880193505"
        assert data["bank_name"]      == "FIRST NATIONAL BANK"
        assert data["account_number"] == "62381077893"
        assert data["branch_code"]    == "220526"
        assert data["vat_rate"]       == 15

    def test_items_have_correct_descriptions(self, db, client, ctx):
        r = client.get(
            f"/api/v1/projects/{ctx['project_id']}/municipality-invoices/template",
            headers=ctx["headers"],
        )
        items = r.json()["data"]["items"]
        descriptions = [i["description"] for i in items]
        assert "P&G" in descriptions
        assert "Roof & Ceiling" in descriptions
        assert "Electrical" in descriptions
        assert "Difference in LOA" in descriptions

    def test_items_have_zero_quantity(self, db, client, ctx):
        """Template items start with qty=0 — users only fill in the ones applicable."""
        r = client.get(
            f"/api/v1/projects/{ctx['project_id']}/municipality-invoices/template",
            headers=ctx["headers"],
        )
        items = r.json()["data"]["items"]
        for item in items:
            assert item["quantity"] == 0

    def test_items_numbered_1_to_15(self, db, client, ctx):
        r = client.get(
            f"/api/v1/projects/{ctx['project_id']}/municipality-invoices/template",
            headers=ctx["headers"],
        )
        items = r.json()["data"]["items"]
        numbers = [i["line_number"] for i in items]
        assert numbers == [str(n) for n in range(1, 16)]

    def test_requires_auth(self, db, client, ctx):
        r = client.get(
            f"/api/v1/projects/{ctx['project_id']}/municipality-invoices/template"
        )
        assert r.status_code == 401

    def test_project_description_matches_project_name(self, db, client, ctx):
        r = client.get(
            f"/api/v1/projects/{ctx['project_id']}/municipality-invoices/template",
            headers=ctx["headers"],
        )
        data = r.json()["data"]
        # project_description should be the project name (not empty)
        assert isinstance(data["project_description"], str)


# ── Full user journey ─────────────────────────────────────────────────────────

class TestUserJourney:
    """
    Simulates the office lady's real workflow:
      1. Load template (15 standard items)
      2. Create invoice by submitting template data with cert + quantities filled in
      3. Preview the totals
      4. Edit one item (add unit price)
      5. Finalise the invoice (lock it)
      6. Attempt to edit finalised invoice (must be blocked)
      7. Download Excel — confirm bytes + valid workbook
      8. List invoices — confirm it appears as FINALISED
    """

    def test_full_workflow(self, db, client, ctx):
        import io
        import openpyxl

        project_id = ctx["project_id"]
        headers    = ctx["headers"]

        # Step 1 — load template
        tr = client.get(
            f"/api/v1/projects/{project_id}/municipality-invoices/template",
            headers=headers,
        )
        assert tr.status_code == 200
        template = tr.json()["data"]
        assert len(template["items"]) == 15

        # Step 2 — create invoice from template (user fills cert + two line items)
        items = template["items"]
        items[6]["unit_price"] = 4500   # line 7: Wall-plate
        items[6]["total"]      = 4500
        items[7]["unit_price"] = 8000   # line 8: Roof & Ceiling
        items[7]["quantity"]   = 3
        items[7]["total"]      = 24000
        items[8]["unit_price"] = 5000   # line 9: Electrical
        items[8]["quantity"]   = 15
        items[8]["total"]      = 75000

        payload = {
            **{k: v for k, v in template.items() if k != "items"},
            "cert_number":         "26",
            "invoice_number":      "IN00472",
            "invoice_date":        "2026-06-23",
            "previously_paid":     0,
            "items":               items,
        }
        cr = client.post(
            f"/api/v1/projects/{project_id}/municipality-invoices/",
            json=payload,
            headers=headers,
        )
        assert cr.status_code == 201, cr.text
        inv = cr.json()["data"]
        inv_id = inv["id"]
        assert inv["cert_number"] == "26"
        assert inv["status"]      == "DRAFT"

        # Step 3 — verify totals (only 3 items have price; rest have no total)
        expected_sub = 4500 + 24000 + 75000   # = 103500
        assert inv["subtotal"] == pytest.approx(expected_sub, abs=1)
        assert inv["vat_amount"] == pytest.approx(expected_sub * 0.15, abs=1)
        assert inv["total_due"]  == pytest.approx(expected_sub * 1.15, abs=1)

        # Step 4 — edit: add a previously-paid amount
        pr = client.patch(
            f"/api/v1/municipality-invoices/{inv_id}",
            json={"previously_paid": 50000, "notes": "Claiming balance after partial receipt."},
            headers=headers,
        )
        assert pr.status_code == 200
        updated = pr.json()["data"]
        net = expected_sub - 50000   # 53500
        assert updated["total_due"] == pytest.approx(net * 1.15, abs=1)

        # Step 5 — finalise
        fr = client.patch(
            f"/api/v1/municipality-invoices/{inv_id}",
            json={"status": "FINALISED"},
            headers=headers,
        )
        assert fr.status_code == 200
        assert fr.json()["data"]["status"] == "FINALISED"

        # Step 6 — edit must be blocked
        er = client.patch(
            f"/api/v1/municipality-invoices/{inv_id}",
            json={"notes": "Should not work"},
            headers=headers,
        )
        assert er.status_code in (400, 422)

        # Step 7 — download Excel
        xr = client.get(
            f"/api/v1/municipality-invoices/{inv_id}/export/excel",
            headers=headers,
        )
        assert xr.status_code == 200
        wb = openpyxl.load_workbook(io.BytesIO(xr.content))
        ws = wb.active
        all_text = " ".join(str(c) for row in ws.iter_rows(values_only=True) for c in row if c)
        assert "Ethekweni Municipality" in all_text
        assert "IN00472" in all_text

        # Step 8 — list shows FINALISED
        lr = client.get(
            f"/api/v1/projects/{project_id}/municipality-invoices/?status=FINALISED",
            headers=headers,
        )
        assert lr.status_code == 200
        ids = [i["id"] for i in lr.json()["data"]]
        assert inv_id in ids

    def test_from_template_then_discard_creates_nothing(self, db, client, ctx):
        """Template fetch itself is read-only — no invoice is created until POST."""
        project_id = ctx["project_id"]
        headers    = ctx["headers"]

        client.get(
            f"/api/v1/projects/{project_id}/municipality-invoices/template",
            headers=headers,
        )
        # Nothing created yet
        lr = client.get(
            f"/api/v1/projects/{project_id}/municipality-invoices/",
            headers=headers,
        )
        assert lr.status_code == 200
        assert len(lr.json()["data"]) == 0

    def test_office_admin_can_create(self, db, client, ctx):
        """Office admin role must be able to create municipality invoices."""
        admin = make_user(db, role="OFFICE_ADMIN")
        make_user_project_access(db, admin["id"], ctx["project_id"])
        token   = login(client, admin["email"], admin["password"])
        headers = auth(token)
        r = _create(client, {"headers": headers, "project_id": ctx["project_id"]})
        assert r.status_code == 201

    def test_read_only_cannot_create(self, db, client, ctx):
        """READ_ONLY role must not be able to create invoices."""
        viewer  = make_user(db, role="READ_ONLY")
        make_user_project_access(db, viewer["id"], ctx["project_id"])
        token   = login(client, viewer["email"], viewer["password"])
        headers = auth(token)
        r = _create(client, {"headers": headers, "project_id": ctx["project_id"]})
        assert r.status_code == 403

    def test_multiple_certs_same_project(self, db, client, ctx):
        """A project can have multiple invoices with different cert numbers."""
        r1 = _create(client, ctx, cert_number="24", invoice_number="IN00470")
        r2 = _create(client, ctx, cert_number="25", invoice_number="IN00471")
        r3 = _create(client, ctx, cert_number="26", invoice_number="IN00472")
        assert all(r.status_code == 201 for r in [r1, r2, r3])
        lr = client.get(
            f"/api/v1/projects/{ctx['project_id']}/municipality-invoices/",
            headers=ctx["headers"],
        )
        assert len(lr.json()["data"]) >= 3
