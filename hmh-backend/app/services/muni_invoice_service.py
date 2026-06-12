"""
Municipality Invoice service.

CRUD + Excel generation matching Cert 26 - Invoice 472.xlsx layout.
No auto-send, no auto-approve. The office lady controls everything.
"""

import io
import uuid
from datetime import date, datetime, timezone
from decimal import Decimal, ROUND_HALF_UP
from typing import Optional

from sqlalchemy.orm import Session, joinedload

from app.core.exceptions import NotFoundError, ValidationError
from app.models.municipality_invoice import MunicipalityInvoice, MunicipalityInvoiceItem
from app.models.project import Project


# ── Number generation ─────────────────────────────────────────────────────────

def _next_invoice_number(db: Session) -> str:
    count = db.query(MunicipalityInvoice).count()
    return f"IN{count + 1:05d}"


# ── Helpers ───────────────────────────────────────────────────────────────────

def _get_or_404(db: Session, invoice_id: uuid.UUID) -> MunicipalityInvoice:
    inv = (
        db.query(MunicipalityInvoice)
        .options(joinedload(MunicipalityInvoice.items))
        .filter(MunicipalityInvoice.id == invoice_id)
        .first()
    )
    if not inv:
        raise NotFoundError(f"Municipality invoice {invoice_id} not found.")
    return inv


def _recalc(inv: MunicipalityInvoice) -> None:
    """Recompute subtotal, vat_amount, total_due from current items."""
    subtotal = sum(
        (item.total or Decimal("0")) for item in inv.items
    )
    inv.subtotal = subtotal.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    net = inv.subtotal - (inv.previously_paid or Decimal("0"))
    if net < 0:
        net = Decimal("0")
    vat_rate = (inv.vat_rate or Decimal("15")) / Decimal("100")
    inv.vat_amount = (net * vat_rate).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    inv.total_due  = (net + inv.vat_amount).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


# ── CRUD ──────────────────────────────────────────────────────────────────────

def list_invoices(
    db: Session,
    project_id: uuid.UUID,
    status: Optional[str] = None,
) -> list[MunicipalityInvoice]:
    if not db.get(Project, project_id):
        raise NotFoundError(f"Project {project_id} not found.")
    q = (
        db.query(MunicipalityInvoice)
        .options(joinedload(MunicipalityInvoice.items))
        .filter(MunicipalityInvoice.project_id == project_id)
    )
    if status:
        q = q.filter(MunicipalityInvoice.status == status)
    return q.order_by(MunicipalityInvoice.created_at.desc()).all()


def get_invoice(db: Session, invoice_id: uuid.UUID) -> MunicipalityInvoice:
    return _get_or_404(db, invoice_id)


def create_invoice(
    db: Session,
    project_id: uuid.UUID,
    created_by: uuid.UUID,
    invoice_number: Optional[str] = None,
    cert_number: Optional[str] = None,
    invoice_date: Optional[date] = None,
    due_date: Optional[date] = None,
    client_name: str = "Ethekweni Municipality",
    client_vat_no: Optional[str] = None,
    client_address: Optional[str] = None,
    company_email: Optional[str] = None,
    project_description: Optional[str] = None,
    contract_reference: Optional[str] = None,
    previously_paid: Optional[Decimal] = None,
    vat_rate: Decimal = Decimal("15.00"),
    notes: Optional[str] = None,
    bank_name: Optional[str] = None,
    account_number: Optional[str] = None,
    branch_name: Optional[str] = None,
    branch_code: Optional[str] = None,
    items: Optional[list[dict]] = None,
) -> MunicipalityInvoice:
    if not db.get(Project, project_id):
        raise NotFoundError(f"Project {project_id} not found.")

    inv_number = invoice_number or _next_invoice_number(db)
    if db.query(MunicipalityInvoice).filter(MunicipalityInvoice.invoice_number == inv_number).first():
        raise ValidationError(f"Invoice number {inv_number} already exists.")

    now = datetime.now(timezone.utc)
    inv = MunicipalityInvoice(
        invoice_number=inv_number,
        cert_number=cert_number,
        project_id=project_id,
        invoice_date=invoice_date or date.today(),
        due_date=due_date,
        client_name=client_name,
        client_vat_no=client_vat_no,
        client_address=client_address,
        company_email=company_email,
        project_description=project_description,
        contract_reference=contract_reference,
        previously_paid=previously_paid,
        vat_rate=vat_rate,
        subtotal=Decimal("0"),
        vat_amount=Decimal("0"),
        total_due=Decimal("0"),
        notes=notes,
        bank_name=bank_name,
        account_number=account_number,
        branch_name=branch_name,
        branch_code=branch_code,
        status="DRAFT",
        created_by=created_by,
        created_at=now,
        updated_at=now,
    )
    db.add(inv)
    db.flush()

    for idx, item_data in enumerate(items or []):
        qty   = item_data.get("quantity")
        price = item_data.get("unit_price")
        total = item_data.get("total")
        if total is None and qty is not None and price is not None:
            total = (Decimal(str(qty)) * Decimal(str(price))).quantize(Decimal("0.01"))
        db.add(MunicipalityInvoiceItem(
            invoice_id  = inv.id,
            sort_order  = item_data.get("sort_order", idx),
            line_number = item_data.get("line_number"),
            description = item_data.get("description", ""),
            quantity    = Decimal(str(qty))    if qty    is not None else None,
            unit_price  = Decimal(str(price))  if price  is not None else None,
            total       = Decimal(str(total))  if total  is not None else None,
            disc_pct    = Decimal(str(item_data["disc_pct"])) if item_data.get("disc_pct") is not None else None,
            comments    = item_data.get("comments"),
            created_at  = now,
        ))

    db.flush()
    db.refresh(inv)
    _recalc(inv)
    db.commit()
    db.refresh(inv)
    return inv


def update_invoice(
    db: Session,
    invoice_id: uuid.UUID,
    **fields,
) -> MunicipalityInvoice:
    inv = _get_or_404(db, invoice_id)
    if inv.status == "FINALISED":
        raise ValidationError("Cannot edit a finalised invoice. Change status to DRAFT first.")

    simple = [
        "cert_number", "invoice_date", "due_date", "client_name", "client_vat_no",
        "client_address", "company_email", "project_description", "contract_reference",
        "previously_paid", "vat_rate", "notes", "bank_name", "account_number",
        "branch_name", "branch_code", "status",
    ]
    for key in simple:
        if key in fields and fields[key] is not None:
            setattr(inv, key, fields[key])

    # Replace items if provided
    if "items" in fields and fields["items"] is not None:
        for old in list(inv.items):
            db.delete(old)
        db.flush()
        now = datetime.now(timezone.utc)
        for idx, item_data in enumerate(fields["items"]):
            qty   = item_data.get("quantity")
            price = item_data.get("unit_price")
            total = item_data.get("total")
            if total is None and qty is not None and price is not None:
                total = (Decimal(str(qty)) * Decimal(str(price))).quantize(Decimal("0.01"))
            db.add(MunicipalityInvoiceItem(
                invoice_id  = inv.id,
                sort_order  = item_data.get("sort_order", idx),
                line_number = item_data.get("line_number"),
                description = item_data.get("description", ""),
                quantity    = Decimal(str(qty))   if qty   is not None else None,
                unit_price  = Decimal(str(price)) if price is not None else None,
                total       = Decimal(str(total)) if total is not None else None,
                disc_pct    = Decimal(str(item_data["disc_pct"])) if item_data.get("disc_pct") is not None else None,
                comments    = item_data.get("comments"),
                created_at  = now,
            ))
        db.flush()
        db.refresh(inv)

    _recalc(inv)
    inv.updated_at = datetime.now(timezone.utc)
    db.commit()
    db.refresh(inv)
    return inv


def delete_invoice(db: Session, invoice_id: uuid.UUID) -> None:
    inv = _get_or_404(db, invoice_id)
    if inv.status == "FINALISED":
        raise ValidationError("Cannot delete a finalised invoice.")
    db.delete(inv)
    db.commit()


# ── Excel export ──────────────────────────────────────────────────────────────

def export_excel(db: Session, invoice_id: uuid.UUID) -> bytes:
    """
    Generate an Excel file matching the Cert 26 - Invoice 472.xlsx layout.
    Returns raw bytes (application/vnd.openxmlformats-officedocument.spreadsheetml.sheet).
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ValidationError("openpyxl is required for Excel export. Run: pip install openpyxl")

    def _add_border(ws_inner, row, col, top=None, bottom=None, left=None, right=None):
        c = ws_inner.cell(row=row, column=col)
        b = c.border
        def _s(existing, new):
            if new:
                return new
            if existing and existing.border_style:
                return existing
            return Side()
        c.border = Border(
            top=_s(b.top if b else None, top),
            bottom=_s(b.bottom if b else None, bottom),
            left=_s(b.left if b else None, left),
            right=_s(b.right if b else None, right),
        )

    inv = _get_or_404(db, invoice_id)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Cert {inv.cert_number or ''} - inv {inv.invoice_number}"

    # ── Style constants (matching original cell-by-cell) ──────────────────────
    MED   = Side(style="medium")
    DBL   = Side(style="double")
    MONEY = '"R"\\ #,##0.00;[Red]"R"\\ #,##0.00'

    F_TITLE  = Font(name="Arial",    size=10, bold=True)
    F_HDR8B  = Font(name="Arial",    size=8,  bold=True)
    F_HDR8   = Font(name="Arial",    size=8,  bold=False)
    F_CAL11B = Font(name="Calibri",  size=11, bold=True,  color="FF000000")
    F_CAL11  = Font(name="Calibri",  size=11, bold=False, color="FF000000")
    F_CAL9   = Font(name="Calibri",  size=9,  bold=False, color="FF000000")
    F_CAL12R = Font(name="Calibri",  size=12, bold=False, color="FFFF0000")
    F_BANK   = Font(name="Copperplate Gothic Bold", size=9)
    F_TOT    = Font(name="Arial",    size=10, bold=True)
    F_DUE    = Font(name="Arial",    size=10, bold=True,  color="FFFF0000")
    F_SIG    = Font(name="Arial",    size=12, bold=False)

    # ── Column widths (exact match to original) ───────────────────────────────
    for col, width in {
        1: 6.44, 2: 0.55, 3: 30.22, 4: 7.33, 5: 0.55,
        6: 6.78, 7: 11.33, 8: 7.66, 9: 12.55, 10: 14.78,
    }.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    # ── Row heights for fixed header section ──────────────────────────────────
    for r, h in {
        1: 15.0, 2: 17.4, 3: 13.2, 4: 14.4, 7: 14.4,
        8: 14.4, 9: 14.4, 10: 15.0, 15: 10.8, 16: 13.8,
        17: 10.8, 18: 22.2, 19: 3.0,
    }.items():
        ws.row_dimensions[r].height = h

    # ── ROW 1: TAX INVOICE ────────────────────────────────────────────────────
    ws.merge_cells("H1:J1")
    c = ws["H1"]
    c.value = "TAX INVOICE"
    c.font  = F_TITLE
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = Border(right=MED, top=MED, bottom=MED)

    # ── ROWS 2–3: Invoice date ────────────────────────────────────────────────
    ws.merge_cells("H2:J3")
    ws.merge_cells("A3:G3")
    c = ws["H2"]
    c.value  = inv.invoice_date if inv.invoice_date else ""
    c.font   = F_HDR8B
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.number_format = "d-mmm-yy"

    # ── ROW 4: spacer merges ──────────────────────────────────────────────────
    ws.merge_cells("F4:G4")

    # ── ROWS 5–6: Page number ─────────────────────────────────────────────────
    ws.merge_cells("F5:G5")
    ws.merge_cells("F6:G6")
    ws.merge_cells("H5:J6")
    c = ws["H5"]
    c.value = "Page                                                     1"
    c.font  = F_HDR8B
    c.alignment = Alignment(horizontal="left", vertical="center")

    # ── ROWS 8–10: Document number ────────────────────────────────────────────
    ws.merge_cells("H8:J10")
    c = ws["H8"]
    c.value = f"Document Number                   {inv.invoice_number}"
    c.font  = F_HDR8B
    c.alignment = Alignment(vertical="center")

    # ── ROW 10: e-mail ────────────────────────────────────────────────────────
    ws.merge_cells("A10:C10")
    c = ws["A10"]
    c.value = "e-mail:"
    c.font  = F_HDR8
    c.alignment = Alignment(horizontal="right")
    c = ws["D10"]
    c.value = inv.company_email or ""
    c.alignment = Alignment(horizontal="left")

    # ── ROW 12: Deliver To ────────────────────────────────────────────────────
    c = ws["H12"]
    c.value = "   Deliver To:"
    c.font  = F_HDR8B
    c.alignment = Alignment(horizontal="left")

    # ── ROW 13: Client VAT + name ─────────────────────────────────────────────
    ws.merge_cells("H13:J13")
    if inv.client_vat_no:
        c = ws["C13"]
        c.value = f"Client Vat No {inv.client_vat_no}"
        c.font  = F_HDR8
    c = ws["H13"]
    c.value = inv.client_name or "Ethekweni Municipality"
    c.font  = F_HDR8
    c.alignment = Alignment(horizontal="right")

    # ── ROWS 14–15: Client address ────────────────────────────────────────────
    ws.merge_cells("H14:J14")
    ws.merge_cells("H15:J15")
    addr = inv.client_address or "PO BOX 828,Durban\n4001"
    addr_lines = addr.split("\n")
    for i, (row, addr_line) in enumerate(zip([14, 15], addr_lines[:2])):
        c = ws.cell(row=row, column=8, value=addr_line)
        c.font  = F_HDR8
        c.alignment = Alignment(horizontal="right")

    # ── ROW 16: Reference column headers ─────────────────────────────────────
    ws.merge_cells("C16:D16")
    ws.merge_cells("F16:G16")
    for col, val, align in [
        (1, "Account",        None),
        (3, "Your Reference", "center"),
        (6, "Tax Reference",  "right"),
        (8, "Sales Code",     "right"),
    ]:
        c = ws.cell(row=16, column=col, value=val)
        c.font = F_HDR8B
        if align:
            c.alignment = Alignment(horizontal=align)

    # ── ROW 17: Reference values ──────────────────────────────────────────────
    ws.merge_cells("C17:D17")
    c = ws["F17"]
    c.value = "N/A"
    c.font  = F_HDR8
    c.alignment = Alignment(horizontal="center")

    # ── ROW 18: Column headers with medium top border ─────────────────────────
    for col, val, align in [
        (1, "Code",        None),
        (3, "Description", None),
        (4, "Quantity",    "center"),
        (6, "Unit Price",  "left"),
        (7, "Total",       "right"),
        (8, "Disc%",       "center"),
        (9, "Tax",         None),
    ]:
        c = ws.cell(row=18, column=col, value=val)
        c.font = F_HDR8B
        if align:
            c.alignment = Alignment(horizontal=align, wrap_text=(col == 6))

    # ── ROW 19: 3pt spacer (already set above) ────────────────────────────────

    # ── ROWS 20–26: Project / cert block ─────────────────────────────────────
    for r in range(20, 27):
        ws.row_dimensions[r].height = 12.0

    c = ws["C21"]
    c.value = inv.project_description or ""
    c.font  = F_CAL11B
    c.alignment = Alignment(horizontal="left")

    c = ws["C23"]
    c.value = inv.contract_reference or ""
    c.font  = F_CAL11B
    c.alignment = Alignment(horizontal="left")

    cert_label = f"(to pay cert {inv.cert_number})" if inv.cert_number else ""
    c = ws["C24"]
    c.value = cert_label
    c.font  = F_CAL11B
    c.alignment = Alignment(horizontal="left")

    sub_val = float(inv.subtotal or 0)
    c = ws["G24"]
    c.value = sub_val
    c.font  = F_TITLE
    c.alignment = Alignment(horizontal="right")
    c.number_format = MONEY

    c = ws["J24"]
    c.value = sub_val
    c.font  = F_TITLE
    c.alignment = Alignment(horizontal="right")
    c.number_format = MONEY

    c = ws["C26"]
    c.value = "DESCRIPTION "
    c.font  = F_CAL11B
    c.alignment = Alignment(horizontal="left")

    # ── ROWS 27+: Line items ─────────────────────────────────────────────────
    ITEM_START = 27
    items = sorted(inv.items, key=lambda x: x.sort_order)
    for idx, item in enumerate(items):
        r = ITEM_START + idx
        ws.row_dimensions[r].height = 12.0

        # Col A: line number — Calibri 11, center/center
        line_num = int(item.line_number) if item.line_number and str(item.line_number).isdigit() else (idx + 1)
        c = ws.cell(row=r, column=1, value=line_num)
        c.font = F_CAL11
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = Border(left=MED)

        # Col C: description — Calibri 9, left/center, wrap
        c = ws.cell(row=r, column=3, value=item.description or "")
        c.font = F_CAL9
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        # Col D: quantity — Calibri 11, center/center
        if item.quantity is not None:
            c = ws.cell(row=r, column=4, value=float(item.quantity))
            c.font = F_CAL11
            c.alignment = Alignment(horizontal="center", vertical="center")

        # Col F: unit price
        if item.unit_price is not None:
            ws.cell(row=r, column=6, value=float(item.unit_price))

        # Col G: total
        if item.total is not None:
            ws.cell(row=r, column=7, value=float(item.total))

        # Col H: disc%
        if item.disc_pct is not None:
            ws.cell(row=r, column=8, value=float(item.disc_pct))

        # Right border on J (outer box — will also be covered by _apply_all_borders)
        ws.cell(row=r, column=10).border = Border(right=MED)

    items_end = ITEM_START + len(items) - 1 if items else ITEM_START - 1

    # ── Notes ─────────────────────────────────────────────────────────────────
    notes_row = items_end + 2
    if inv.notes:
        c = ws.cell(row=notes_row, column=3, value=inv.notes)
        c.font = F_CAL12R
        c.alignment = Alignment(horizontal="left")

    # ── Banking details (Copperplate Gothic Bold 9) ────────────────────────────
    bank_start = notes_row + 5
    if inv.bank_name:
        for i, line in enumerate([
            f"BANKING DETAILS : {inv.bank_name.upper()}",
            f"ACCOUNT NUMBER : {inv.account_number or ''}",
            f"BRANCH  : {inv.branch_name or ''}",
            f"BRANCH CODE: {inv.branch_code or ''}",
        ]):
            c = ws.cell(row=bank_start + i, column=3, value=line)
            c.font = F_BANK
        bank_end = bank_start + 3
    else:
        bank_end = bank_start - 1

    # ── Totals section (rows T through T+7) ───────────────────────────────────
    T = bank_end + 5

    sub  = float(inv.subtotal or 0)
    prev = float(inv.previously_paid or 0)
    net  = max(0.0, sub - prev)
    vat  = float(inv.vat_amount or 0)
    due  = float(inv.total_due or 0)

    ws.row_dimensions[T].height = 14.4
    c = ws.cell(row=T, column=9, value="Sub Total")
    c.font = F_TOT; c.alignment = Alignment(vertical="center")
    c = ws.cell(row=T, column=10, value=sub)
    c.font = F_TOT; c.alignment = Alignment(vertical="center"); c.number_format = MONEY

    ws.row_dimensions[T + 1].height = 15.6
    c = ws.cell(row=T + 1, column=1, value="Signature: "); c.font = F_SIG
    c = ws.cell(row=T + 1, column=10, value=0); c.font = F_TOT; c.number_format = MONEY

    ws.row_dimensions[T + 2].height = 13.2
    c = ws.cell(row=T + 2, column=9, value="Sub Total"); c.font = F_TOT
    c = ws.cell(row=T + 2, column=10, value=sub)
    c.font = F_TOT; c.alignment = Alignment(horizontal="right"); c.number_format = MONEY

    ws.row_dimensions[T + 3].height = 14.4

    ws.row_dimensions[T + 4].height = 13.2
    c = ws.cell(row=T + 4, column=9, value="Less Previously paid")
    c.font = F_TOT; c.alignment = Alignment(horizontal="right")
    c = ws.cell(row=T + 4, column=10, value=prev)
    c.font = F_TOT; c.alignment = Alignment(vertical="center", wrap_text=True); c.number_format = MONEY

    ws.row_dimensions[T + 5].height = 13.2
    c = ws.cell(row=T + 5, column=9, value="Sub Total"); c.font = F_TOT
    c = ws.cell(row=T + 5, column=10, value=net)
    c.font = F_TOT; c.alignment = Alignment(horizontal="right"); c.number_format = MONEY

    ws.row_dimensions[T + 6].height = 13.2
    c = ws.cell(row=T + 6, column=9, value=f"Vat @{float(inv.vat_rate or 15):.0f}%")
    c.font = F_TOT; c.alignment = Alignment(vertical="center")
    c = ws.cell(row=T + 6, column=10, value=vat)
    c.font = F_TOT; c.alignment = Alignment(vertical="center"); c.number_format = MONEY

    ws.row_dimensions[T + 7].height = 13.8
    c = ws.cell(row=T + 7, column=9, value="Now Due "); c.font = F_TOT
    c = ws.cell(row=T + 7, column=10, value=due); c.font = F_DUE; c.number_format = MONEY

    # ── Apply all outer-box + section borders (must come after all values set) ──
    BOTTOM_ROW = T + 8

    def _ab(row, col, top=None, bottom=None, left=None, right=None):
        _add_border(ws, row, col, top=top, bottom=bottom, left=left, right=right)

    # Full outer left/right on every row
    for r in range(1, BOTTOM_ROW + 1):
        _ab(r, 1,  left=MED)
        _ab(r, 10, right=MED)

    # Top outer border row 1
    for col in range(1, 11):
        _ab(1, col, top=MED)

    # Bottom outer border BOTTOM_ROW
    for col in range(1, 11):
        _ab(BOTTOM_ROW, col, bottom=MED)
    _ab(BOTTOM_ROW, 8, left=MED)

    # Vertical divider: G right rows 1-10
    for r in range(1, 11):
        _ab(r, 7, right=MED)

    # H1:J1 sub-box
    _ab(1, 8, top=MED, bottom=MED, right=MED)
    for col in (9, 10):
        _ab(1, col, top=MED, bottom=MED)

    # H2:J3 date area
    _ab(2, 8, top=MED, right=MED)
    _ab(2, 9, top=MED); _ab(2, 10, top=MED)

    # H5 right (page area)
    _ab(5, 8, right=MED)

    # H8 right (document number)
    _ab(8, 8, right=MED)

    # Row 10 bottom (close left-header section)
    for col in range(1, 8):
        _ab(10, col, bottom=MED)

    # H11:J11 top (start Deliver-To sub-box)
    for col in range(8, 11):
        _ab(11, col, top=MED)
    _ab(11, 8, left=MED)

    # H col left rows 12-15
    for r in range(12, 16):
        _ab(r, 8, left=MED)

    # H15:J15 bottom (close client box)
    for col in range(8, 11):
        _ab(15, col, bottom=MED)
    _ab(15, 8, left=MED, right=MED)

    # Row 16 full top+bottom
    for col in range(1, 11):
        _ab(16, col, top=MED, bottom=MED)

    # Rows 18-20 top separators
    for col in range(1, 11):
        _ab(18, col, top=MED)
        _ab(19, col, top=MED)
        _ab(20, col, top=MED)

    # Totals H col left (rows T through T+7)
    for r in range(T, T + 8):
        _ab(r, 8, left=MED)

    # Totals top separator row T
    for col in range(1, 11):
        _ab(T, col, top=MED)

    # Now Due double-bottom on I and J (must come after outer-box pass)
    _ab(T + 7, 8, left=MED)
    _ab(T + 7, 9, bottom=DBL)
    _ab(T + 7, 10, bottom=DBL)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
