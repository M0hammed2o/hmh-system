"""
Export Service — Phase 5.
Generates PDF and Excel management reports from analytics data.
"""
from io import BytesIO
from typing import Any


# ── Excel ─────────────────────────────────────────────────────────────────────

def _wb_header_style(ws, headers: list[str], col_widths: list[int] | None = None):
    """Write a bold header row with background fill."""
    from openpyxl.styles import Alignment, Font, PatternFill

    fill = PatternFill(fill_type="solid", fgColor="1E3A5F")
    font = Font(bold=True, color="FFFFFF", size=11)

    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = font
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    if col_widths:
        from openpyxl.utils import get_column_letter
        for i, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w


def _rnd(v) -> str:
    """Format a monetary value for Excel cells."""
    try:
        return f"R {float(v):,.2f}"
    except (TypeError, ValueError):
        return str(v) if v is not None else ""


# ── Supplier Spend Excel ──────────────────────────────────────────────────────

def export_supplier_spend_excel(data: dict) -> BytesIO:
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Supplier Spend"
    ws.row_dimensions[1].height = 20

    headers = ["Supplier", "Code", "Total Spend", "Spend This Year", "Spend This Month", "PO Count"]
    widths   = [30, 15, 20, 20, 20, 12]
    _wb_header_style(ws, headers, widths)

    for r, sup in enumerate(data.get("top_suppliers", []), start=2):
        ws.cell(r, 1, sup["supplier_name"])
        ws.cell(r, 2, sup["supplier_code"] or "")
        ws.cell(r, 3, _rnd(sup["total_spend"]))
        ws.cell(r, 4, _rnd(sup["spend_this_year"]))
        ws.cell(r, 5, _rnd(sup["spend_this_month"]))
        ws.cell(r, 6, sup["po_count"])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── Outstanding Orders Excel ──────────────────────────────────────────────────

def export_outstanding_orders_excel(projects: list[dict]) -> BytesIO:
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Outstanding Orders"

    headers = ["Project", "Code", "Open POs", "Pending Deliveries", "Total Spend", "BOQ Budget", "Variance"]
    widths   = [30, 12, 12, 18, 20, 20, 20]
    _wb_header_style(ws, headers, widths)

    for r, p in enumerate(projects, start=2):
        ws.cell(r, 1, p["project_name"])
        ws.cell(r, 2, p["project_code"])
        ws.cell(r, 3, p["open_pos"])
        ws.cell(r, 4, p["outstanding_deliveries"])
        ws.cell(r, 5, _rnd(p["total_spend"]))
        ws.cell(r, 6, _rnd(p["boq_budget"]))
        ws.cell(r, 7, _rnd(p["budget_variance"]))

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── Savings Report Excel ──────────────────────────────────────────────────────

def export_savings_excel(savings: list[dict]) -> BytesIO:
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Procurement Savings"

    headers = ["Project", "Code", "BOQ Budget", "Actual Spend", "Variance", "Variance %", "Status"]
    widths   = [30, 12, 20, 20, 20, 14, 14]
    _wb_header_style(ws, headers, widths)

    for r, s in enumerate(savings, start=2):
        ws.cell(r, 1, s["project_name"])
        ws.cell(r, 2, s["project_code"])
        ws.cell(r, 3, _rnd(s["boq_budget"]))
        ws.cell(r, 4, _rnd(s["actual_spend"]))
        ws.cell(r, 5, _rnd(s["variance"]))
        ws.cell(r, 6, f"{s['variance_pct']:.2f}%")
        ws.cell(r, 7, "Under Budget" if s["status"] == "under_budget" else "Over Budget")

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── Reconciliation Report Excel ───────────────────────────────────────────────

def export_reconciliation_excel(recons: list[dict]) -> BytesIO:
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reconciliations"

    headers = ["Rec #", "PO #", "Supplier", "Status", "Invoice #", "Created"]
    widths   = [20, 20, 30, 22, 20, 18]
    _wb_header_style(ws, headers, widths)

    for r, rec in enumerate(recons, start=2):
        po = rec.get("po") or {}
        inv = rec.get("invoice") or {}
        ws.cell(r, 1, rec.get("reconciliation_number", ""))
        ws.cell(r, 2, po.get("po_number", ""))
        ws.cell(r, 3, po.get("supplier_name", ""))
        ws.cell(r, 4, rec.get("status", ""))
        ws.cell(r, 5, inv.get("invoice_number", ""))
        ws.cell(r, 6, rec.get("created_at", "")[:10])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── PDF helpers ───────────────────────────────────────────────────────────────

def _pdf_table_style(col_widths, data: list[list], title: str) -> BytesIO:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=30, rightMargin=30, topMargin=30, bottomMargin=30)
    styles = getSampleStyleSheet()

    header_bg = colors.HexColor("#1E3A5F")
    alt_row   = colors.HexColor("#F0F4F8")
    white     = colors.white

    style = TableStyle([
        ("BACKGROUND",  (0, 0), (-1, 0),  header_bg),
        ("TEXTCOLOR",   (0, 0), (-1, 0),  white),
        ("FONTNAME",    (0, 0), (-1, 0),  "Helvetica-Bold"),
        ("FONTSIZE",    (0, 0), (-1, 0),  9),
        ("ALIGN",       (0, 0), (-1, 0),  "CENTER"),
        ("VALIGN",      (0, 0), (-1, -1), "MIDDLE"),
        ("ROWHEIGHT",   (0, 0), (-1, -1), 18),
        ("FONTNAME",    (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE",    (0, 1), (-1, -1), 8),
        ("GRID",        (0, 0), (-1, -1), 0.25, colors.HexColor("#CBD5E0")),
        ("LEFTPADDING",  (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
    ])
    # Alternate row shading
    for i in range(1, len(data)):
        if i % 2 == 0:
            style.add("BACKGROUND", (0, i), (-1, i), alt_row)

    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(style)

    story = [
        Paragraph(f"<b>{title}</b>", styles["Title"]),
        Spacer(1, 12),
        table,
    ]
    doc.build(story)
    buf.seek(0)
    return buf


# ── Supplier Spend PDF ────────────────────────────────────────────────────────

def export_supplier_spend_pdf(data: dict) -> BytesIO:
    headers = [["Supplier", "Code", "Total Spend", "This Year", "This Month", "POs"]]
    rows = [
        [
            s["supplier_name"], s["supplier_code"] or "",
            _rnd(s["total_spend"]), _rnd(s["spend_this_year"]),
            _rnd(s["spend_this_month"]), str(s["po_count"]),
        ]
        for s in data.get("top_suppliers", [])
    ]
    return _pdf_table_style(
        [130, 60, 90, 90, 90, 40],
        headers + rows,
        "Supplier Spend Report",
    )


# ── Outstanding Orders PDF ────────────────────────────────────────────────────

def export_outstanding_orders_pdf(projects: list[dict]) -> BytesIO:
    headers = [["Project", "Code", "Open POs", "Pend. Deliveries", "Total Spend", "BOQ Budget", "Variance"]]
    rows = [
        [
            p["project_name"], p["project_code"],
            str(p["open_pos"]), str(p["outstanding_deliveries"]),
            _rnd(p["total_spend"]), _rnd(p["boq_budget"]), _rnd(p["budget_variance"]),
        ]
        for p in projects
    ]
    return _pdf_table_style(
        [130, 50, 60, 80, 90, 90, 90],
        headers + rows,
        "Outstanding Orders Report",
    )


# ── Savings PDF ───────────────────────────────────────────────────────────────

def export_savings_pdf(savings: list[dict]) -> BytesIO:
    headers = [["Project", "Code", "BOQ Budget", "Actual Spend", "Variance", "Variance %", "Status"]]
    rows = [
        [
            s["project_name"], s["project_code"],
            _rnd(s["boq_budget"]), _rnd(s["actual_spend"]),
            _rnd(s["variance"]), f"{s['variance_pct']:.2f}%",
            "Under Budget" if s["status"] == "under_budget" else "Over Budget",
        ]
        for s in savings
    ]
    return _pdf_table_style(
        [120, 50, 90, 90, 90, 60, 70],
        headers + rows,
        "Procurement Savings Report",
    )


# ── Reconciliation PDF ────────────────────────────────────────────────────────

def export_reconciliation_pdf(recons: list[dict]) -> BytesIO:
    headers = [["Rec #", "PO #", "Supplier", "Status", "Invoice #", "Created"]]
    rows = [
        [
            rec.get("reconciliation_number", ""),
            (rec.get("po") or {}).get("po_number", ""),
            (rec.get("po") or {}).get("supplier_name", ""),
            rec.get("status", ""),
            (rec.get("invoice") or {}).get("invoice_number", ""),
            (rec.get("created_at") or "")[:10],
        ]
        for rec in recons
    ]
    return _pdf_table_style(
        [90, 90, 130, 100, 90, 70],
        headers + rows,
        "Reconciliation Report",
    )
